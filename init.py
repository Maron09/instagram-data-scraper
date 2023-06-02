from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import json
import uuid
import requests
import os
import pandas as pd
import openpyxl

def get_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('disable-infobars')
    options.add_argument("start-maximized")
    options.add_argument("disable-dev-shm-usage")
    options.add_argument("no-sandbox")
    options.add_experimental_option("excludeSwitches", ['enable-automation'])
    options.add_argument("disable-blink-features=AutomationControlled")
    
    driver = webdriver.Chrome(options=options)
    driver.get("https://www.instagram.com/accounts/login/")
    return driver

num_elements_to_scrape = 5000
session_id = input("Enter SessionID: ")
tag = input('Enter Search Key: ')

def main():
    try:
        driver = get_driver()
        driver.implicitly_wait(10)
        driver.add_cookie({'name': 'sessionid', 'value': session_id})
        driver.refresh()
        url = f'https://www.instagram.com/explore/tags/{tag}/'
        driver.get(url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div._aabd._aa8k._al3l a')))

        last_height = driver.execute_script("return document.body.scrollHeight")
        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            new_height = driver.execute_script("return document.body.scrollHeight")

            # Check if the total number of elements is reached
            posts = driver.find_elements(By.CSS_SELECTOR, 'div._aabd._aa8k._al3l a')
            if len(posts) >= num_elements_to_scrape:
                break
            if new_height == last_height:
                break

            last_height = new_height

        posts = posts[:num_elements_to_scrape]
        scraped_data = []
        for post in posts:
            links = post.get_attribute('href')

            scraped_data.append(links)
        print(scraped_data)
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    finally:
        driver.close()

    def get_info():
        output_folder = 'output'
        output_file = 'output.xlsx'

        # Create the output folder if it doesn't exist
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Create a new Excel file or load an existing one
        if not os.path.exists(output_file):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.append(["Filename", "Name", "Username", "Phone Number"])
        else:
            workbook = openpyxl.load_workbook(output_file)
            worksheet = workbook.active

        for data in scraped_data:
            content = requests.get(data).text
            pattern = r'<script type="application/ld\+json" nonce="[^"]*">(.*?)</script>'
            matches = re.findall(pattern, content, re.DOTALL)

            if matches:
                script_content = matches[0]

                # Convert the script content to a dictionary
                json_data = json.loads(script_content)

                # Generate a random name for the JSON file
                file_name = str(uuid.uuid4()) + '.json'
                output_path = os.path.join(output_folder, file_name)

                # Save the dictionary as a JSON file with the random name
                with open(output_path, 'w') as json_file:
                    json.dump(json_data, json_file, indent=4)

                print(f"JSON file saved in folder: {output_folder}, with filename: {file_name}")
            else:
                print("Script tag not found or invalid format")

            for file_name in os.listdir(output_folder):
                # Check if the file is a JSON file
                if file_name.endswith('.json'):
                    # Construct the file paths
                    json_file_path = os.path.join(output_folder, file_name)
                    text_file_path = os.path.splitext(json_file_path)[0] + '.txt'

                    # Read the JSON file
                    with open(json_file_path, 'r') as json_file:
                        json_data = json.load(json_file)

                    # Write the JSON content as text to a new text file
                    with open(text_file_path, 'w') as text_file:
                        text_file.write(json.dumps(json_data))

                    print(f"Converted JSON file '{file_name}' to text file '{os.path.basename(text_file_path)}'")

                    name_pattern = r'"name": "(.*?)"'
                    alternate_name_pattern = r'"alternateName": "(.*?)"'
                    phone_number_pattern = r'CALL (\d{11})'
                    result = []
                    try:
                        with open(text_file_path, "r") as file:
                            contents = file.read()

                            match_name = re.search(name_pattern, contents)
                            name = match_name.group(1) if match_name else ""

                            match_alternate_name = re.search(alternate_name_pattern, contents)
                            alternate_name = match_alternate_name.group(1) if match_alternate_name else ""

                            match_phone_number = re.search(phone_number_pattern, contents)
                            phone_number = match_phone_number.group(1) if match_phone_number else ""

                            result.append({
                                "Filename": file_name,
                                "Name": name,
                                "Username": alternate_name,
                                "Phone Number": phone_number
                            })

                    except (FileNotFoundError, IOError):
                        print(f"Error reading file: {file_name}")

                    if result:
                        for entry in result:
                            worksheet.append([
                                entry["Filename"],
                                entry["Name"],
                                entry["Username"],
                                entry["Phone Number"]
                            ])

        # Remove duplicates from the worksheet based on the "Name" column
        df = pd.DataFrame(worksheet.values)
        df.columns = df.iloc[0]
        df = df[1:]
        df.drop_duplicates(subset="Name", keep="first", inplace=True)

        # Clear the existing worksheet data
        worksheet.delete_rows(2, worksheet.max_row)

        # Append the updated data to the worksheet
        for row in df.values:
            worksheet.append(row.tolist())

        # Save the Excel file
        workbook.save(output_file)
        print(f"Data appended to {output_file}.")

    get_info()


if __name__ == "__main__":
    main()
